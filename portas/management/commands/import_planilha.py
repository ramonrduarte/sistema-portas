from django.core.management.base import BaseCommand
from django.conf import settings
from pathlib import Path
import openpyxl

from portas.models import (
    Perfil,
    PerfilPuxador,
    Puxador,
    Divisor,
    VidroBase,
)


def limpa_codigo(codigo):
    if codigo is None:
        return None
    codigo = str(codigo).strip()
    return codigo


class Command(BaseCommand):
    help = "Importa dados da planilha 'Calculo de Portas 2025.xlsx' para o banco de dados"

    def handle(self, *args, **options):
        base_dir = Path(settings.BASE_DIR)
        arquivo = base_dir / "dados" / "Calculo de Portas 2025.xlsx"

        self.stdout.write(self.style.NOTICE(f"Lendo arquivo: {arquivo}"))

        wb = openpyxl.load_workbook(arquivo, data_only=True)

        self.importar_perfil(wb["PERFIl"])
        self.importar_perfil_puxador(wb["PERFIL_PUXADOR"])
        self.importar_puxador(wb["PUXADOR"])
        self.importar_divisores(wb["DIVISORES"])
        self.importar_vidros(wb["BASE_VIDROS"])

        self.stdout.write(self.style.SUCCESS("Importação concluída!"))

    def importar_perfil(self, sheet):
        self.stdout.write("Importando PERFIl...")
        for linha in sheet.iter_rows(min_row=2, values_only=True):
            codigo, descricao, preco = linha[:3]
            if not codigo or not descricao or preco is None:
                continue
            Perfil.objects.update_or_create(
                codigo=limpa_codigo(codigo),
                defaults={"descricao": descricao, "preco": preco},
            )

    def importar_perfil_puxador(self, sheet):
        self.stdout.write("Importando PERFIL_PUXADOR...")
        for linha in sheet.iter_rows(min_row=2, values_only=True):
            codigo, descricao, preco = linha[:3]
            if not codigo or not descricao or preco is None:
                continue
            PerfilPuxador.objects.update_or_create(
                codigo=limpa_codigo(codigo),
                defaults={"descricao": descricao, "preco": preco},
            )

    def importar_puxador(self, sheet):
        self.stdout.write("Importando PUXADOR...")
        for linha in sheet.iter_rows(min_row=2, values_only=True):
            codigo, descricao, preco = linha[:3]
            if not codigo or not descricao or preco is None:
                continue
            Puxador.objects.update_or_create(
                codigo=limpa_codigo(codigo),
                defaults={"descricao": descricao, "preco": preco},
            )

    def importar_divisores(self, sheet):
        self.stdout.write("Importando DIVISORES...")
        for linha in sheet.iter_rows(min_row=2, values_only=True):
            codigo, descricao, preco, acabamento, tipo, modelo = linha[:6]
            if not codigo or not descricao or preco is None:
                continue
            Divisor.objects.update_or_create(
                codigo=limpa_codigo(codigo),
                defaults={
                    "descricao": descricao,
                    "preco": preco,
                    "modelo": str(modelo) if modelo is not None else "",
                },
            )

    def importar_vidros(self, sheet):
        self.stdout.write("Importando BASE_VIDROS...")
        for linha in sheet.iter_rows(min_row=2, values_only=True):
            codigo, descricao, preco = linha[:3]
            if not codigo or not descricao or preco is None:
                continue
            VidroBase.objects.update_or_create(
                codigo=limpa_codigo(codigo),
                defaults={"descricao": descricao, "preco": preco},
            )
